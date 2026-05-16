#!/usr/bin/env python3
"""
pdf_to_booking_excel.py
Convert erlebe-fernreisen booking PDFs to a two-sheet Excel:
  Sheet 1 – "DB Import"   → one row per column, maps to the quotations table
  Sheet 2 – "Itinerary"   → day-by-day reference for staff

Usage:
    python pdf_to_booking_excel.py <input.pdf> [output.xlsx]
"""

import sys
import json
from pathlib import Path
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import the converter for superior rule-based parsing
from converter import extract_pdf_text, parse_pdf_rules


# ──────────────────────────────────────────────────────────────
# 2. CLAUDE-POWERED PARSING
# ──────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────
# 2. SHARED HELPERS & STYLES
# ──────────────────────────────────────────────────────────────

def to_date(date_str):
    if not date_str:
        return None
    return datetime.strptime(date_str, "%d.%m.%Y").date()

def guest_display(guests: list) -> str:
    return " & ".join(f"{g['title']}. {g['first_name']} {g['last_name']}" for g in guests)

THIN = Side(style="thin", color="BFBFBF")
def tb():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BLUE      = "1F4E79"
ALT_BLUE  = "DCE6F1"
GREEN     = "375623"
ALT_GREEN = "EBF1DE"

def hdr_style(color_hex):
    return (
        Font(name="Calibri", bold=True, color="FFFFFF", size=10),
        PatternFill("solid", fgColor=color_hex),
    )

DAT_FONT = Font(name="Calibri", size=10)
LBL_FONT = Font(name="Calibri", bold=True, size=10)


# ──────────────────────────────────────────────────────────────
# 4. SHEET 1 — DB IMPORT  (quotations table)
# ──────────────────────────────────────────────────────────────

# (db_column_name, friendly_label, json_key, dtype)
DB_COLUMNS = [
    # Identification
    ("QuotationRef",    "Quotation Ref",        "QuotationRef",    "text"),
    ("PrincipalClient", "Principal Client",      "PrincipalClient", "text"),
    ("PaxName",         "Pax Last Name",         "PaxName",         "text"),
    ("PaxFirstName",    "Pax First Name(s)",     "PaxFirstName",    "text"),
    ("Email",           "Email",                 "Email",           "text"),
    # Pax & Rooms
    ("NumPax",          "No. of Pax",            "NumPax",          "int"),
    ("NumSingles",      "Single Rooms",          "NumSingles",      "int"),
    ("NumDoubles",      "Double Rooms",          "NumDoubles",      "int"),
    ("NumTwins",        "Twin Rooms",            "NumTwins",        "int"),
    ("NumTriples",      "Triple Rooms",          "NumTriples",      "int"),
    # Dates
    ("DateOfArrival",   "Date of Arrival",       "DateOfArrival",   "date"),
    ("DateOfDeparture", "Date of Departure",     "DateOfDeparture", "date"),
    ("StartDate",       "Start Date",            "StartDate",       "date"),
    ("EndDate",         "End Date",              "EndDate",         "date"),
    ("Nights",          "No. of Nights",         "Nights",          "int"),
    # Flights
    ("FlightNo",        "Inbound Flight No.",    "FlightNo",        "text"),
    ("ETA",             "ETA",                   "ETA",             "text"),
    ("PlaceFrom",       "Place From",            "PlaceFrom",       "text"),
    ("FlightNoDept",    "Outbound Flight No.",   "FlightNoDept",    "text"),
    ("ETD",             "ETD",                   "ETD",             "text"),
    ("PlaceTo",         "Place To",              "PlaceTo",         "text"),
    # Inclusions
    ("Guide",           "Guide Included",        "Guide",           "bool"),
    ("EntranceFees",    "Entrance Fees Incl.",   "EntranceFees",    "bool"),
    # Notes
    ("Comment",         "Comment / Notes",       "Comment",         "text"),
]

def build_db_sheet(ws, q: dict):
    ws.title = "DB Import"
    ws.column_dimensions["A"].width = 26   # friendly label
    ws.column_dimensions["B"].width = 22   # db column name
    ws.column_dimensions["C"].width = 40   # value

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Quotations — DB Import"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    hdr_font, hdr_fill = hdr_style(BLUE)
    for col, label in enumerate(["Friendly Label", "DB Column Name", "Value"], start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb()
    ws.row_dimensions[2].height = 18

    # Data rows
    for r, (db_col, label, key, dtype) in enumerate(DB_COLUMNS, start=3):
        raw = q.get(key)

        if   dtype == "date": val = to_date(raw)
        elif dtype == "int":  val = int(raw) if raw is not None else 0
        elif dtype == "bool": val = "Yes" if raw else "No"
        else:                 val = raw if raw is not None else ""

        alt = PatternFill("solid", fgColor=ALT_BLUE) if r % 2 == 0 else None

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LBL_FONT; lc.border = tb()
        lc.alignment = Alignment(vertical="center")
        if alt: lc.fill = alt

        dc = ws.cell(row=r, column=2, value=db_col)
        dc.font = Font(name="Courier New", size=9, color="555555")
        dc.border = tb(); dc.alignment = Alignment(vertical="center")
        if alt: dc.fill = alt

        vc = ws.cell(row=r, column=3, value=val)
        vc.font = DAT_FONT; vc.border = tb()
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        if dtype == "date" and val:
            vc.number_format = "DD/MM/YYYY"
        if alt: vc.fill = alt

        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────────────────────
# 5. SHEET 2 — ITINERARY  (staff reference)
# ──────────────────────────────────────────────────────────────

def build_itinerary_sheet(ws, data: dict):
    ws.title = "Itinerary"
    for col, w in zip("ABCDE", [8, 12, 14, 20, 80]):
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = guest_display(data["guests"])
    ws["A1"].font = Font(name="Calibri", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Summary row
    q = data["quotation"]
    parts = [f"{q[k]}×{lbl}" for k, lbl in
             [("NumDoubles","Double"),("NumTwins","Twin"),("NumSingles","Single"),("NumTriples","Triple")]
             if q.get(k)]
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Pax: {q['NumPax']}   |   Rooms: {', '.join(parts)}   |   Nights: {q['Nights']}"
    ws["A2"].font = Font(name="Calibri", bold=True, size=10, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Column headers
    hdr_font, hdr_fill = hdr_style(GREEN)
    for col, hdr in enumerate(["Sr. No", "Day", "Date", "Destination", "Itinerary"], start=1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb()
    ws.row_dimensions[4].height = 18

    # Day rows
    for idx, entry in enumerate(data["itinerary"]):
        row = 5 + idx
        alt = PatternFill("solid", fgColor=ALT_GREEN) if idx % 2 == 1 else None
        vals = [entry["sr_no"], entry["day_name"], to_date(entry["date"]),
                entry["destination"], entry["description"]]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = DAT_FONT; c.border = tb()
            if alt: c.fill = alt
            if col in (1, 2, 4):
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif col == 3:
                c.number_format = "DD/MM/YYYY"
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 44

    ws.freeze_panes = "A5"


# ──────────────────────────────────────────────────────────────
# 6. BUILD WORKBOOK
# ──────────────────────────────────────────────────────────────

def build_excel(data: dict, output_path: str):
    wb = Workbook()
    build_db_sheet(wb.active, data["quotation"])
    build_itinerary_sheet(wb.create_sheet(), data)
    wb.save(output_path)
    print(f"[OK] Saved: {output_path}")


# ──────────────────────────────────────────────────────────────
# 7. MAIN
# ──────────────────────────────────────────────────────────────


def main():
    if len(sys.argv) < 2:
        print("Usage: python pdf_to_booking_excel.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path    = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) >= 3 else str(Path(pdf_path).with_suffix(".xlsx"))

    print(f"[PDF] Reading: {pdf_path}")
    raw_text = extract_pdf_text(pdf_path)

    print("[PARSE] Parsing with rule-based parser ...")
    data = parse_pdf_rules(raw_text)

    q = data.get('quotation', {})
    print(f"  Client : {q.get('PrincipalClient')}")
    print(f"  Ref    : {q.get('QuotationRef')}")
    print(f"  Pax    : {q.get('NumPax')}")
    print(f"  Arrival: {q.get('DateOfArrival')} -> Dep: {q.get('DateOfDeparture')}")
    print(f"  Days   : {len(data.get('itinerary', []))} itinerary entries")


    print("[BUILD] Building Excel ...")
    build_excel(data, output_path)


if __name__ == "__main__":
    main()
