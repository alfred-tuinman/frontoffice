"""write_ids_to_excel.py — Append DB IDs to the DB Import sheet of a booking Excel."""
import sys
from openpyxl import load_workbook


def write_ids_to_excel(xlsx_path: str, quotations_id: int, itineraries_id: int) -> None:
    """
    Open the Excel file at xlsx_path and append two rows to the 'DB Import' sheet:
      Quotations ID  | Quotations_id  | <value>
      Itineraries ID | itineraries_id | <value>
    Overwrites existing ID rows if already present.
    """
    wb = load_workbook(xlsx_path)

    if "DB Import" not in wb.sheetnames:
        raise ValueError(f"'DB Import' sheet not found in {xlsx_path}")

    ws = wb["DB Import"]

    # Find the last used row in column B (DB Column Name)
    # Remove any existing ID rows first to avoid duplicates on re-save
    rows_to_delete = []
    for row in ws.iter_rows():
        if row[1].value in ("Quotations_id", "itineraries_id"):
            rows_to_delete.append(row[0].row)

    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)

    # Find new last row and append
    last_row = ws.max_row + 1
    ws.cell(row=last_row,     column=1, value="Quotations ID")
    ws.cell(row=last_row,     column=2, value="Quotations_id")
    ws.cell(row=last_row,     column=3, value=quotations_id)

    ws.cell(row=last_row + 1, column=1, value="Itineraries ID")
    ws.cell(row=last_row + 1, column=2, value="itineraries_id")
    ws.cell(row=last_row + 1, column=3, value=itineraries_id)

    wb.save(xlsx_path)


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python write_ids_to_excel.py <xlsx_path> <quotations_id> <itineraries_id>")
        sys.exit(1)
    write_ids_to_excel(sys.argv[1], int(sys.argv[2]), int(sys.argv[3]))
    print(f"✅ IDs written to {sys.argv[1]}")