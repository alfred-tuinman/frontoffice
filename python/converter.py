"""
converter.py — Rule-based PDF parser for erlebe-fernreisen booking PDFs.
No external API required — all parsing done locally via pdfplumber + regex.
"""

import re
from datetime import datetime, timedelta
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Regex patterns ─────────────────────────────────────────────────────────────

_DAYS = r'Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday'

# Date pattern — matches all common international date formats found in booking PDFs.
# Normalisation of U+2010 non-breaking hyphens → ASCII hyphen is done in _normalise_text().
#
# Formats covered (separators may be . / - or space):
#   DD.MM.YYYY  DD-MM-YYYY  DD/MM/YYYY          European / UK / India / Australia
#   DD-Mon-YYYY DD Mon YYYY DD. Mon YYYY         Textual month, any region
#   YYYY-MM-DD  YYYY/MM/DD                       ISO 8601, East Asia
#   MM/DD/YYYY  MM-DD-YYYY                       United States
_DATE_PAT = (
    r'(?:'
    r'\d{4}[\-/]\d{2}[\-/]\d{2}'            # YYYY-MM-DD / YYYY/MM/DD
    r'|\d{2}[.\-/]\d{2}[.\-/]\d{4}'          # DD.MM.YYYY / DD-MM-YYYY / DD/MM/YYYY / MM/DD/YYYY
    r'|\d{1,2}[.\-/\s][A-Za-z]{3,9}[.\-/\s]\d{4}'  # DD Mon YYYY / DD-Mon-YYYY / DD. Month YYYY
    r')'
)

# Booking reference — tries several common label phrasings in order
_BOOKING_REF_PATTERNS = [
    re.compile(r'booking\s+(?:number|no\.?|ref(?:erence)?)\s*[:\-]?\s*(\S+)', re.IGNORECASE),
    re.compile(r'(?:reservation|voucher|confirmation)\s+(?:number|no\.?|ref(?:erence)?)\s*[:\-]?\s*(\S+)', re.IGNORECASE),
    re.compile(r'(?:ref(?:erence)?|ref\.?)\s*[:\-]\s*(\S+)', re.IGNORECASE),
    re.compile(r'(?:buchungsnummer|buchungs-nr\.?)\s*[:\-]?\s*(\S+)', re.IGNORECASE),
]
BOOKING_REF_RE = _BOOKING_REF_PATTERNS[0]  # kept for compat; _search_booking_ref used internally

# Guest line — four increasingly-relaxed patterns tried in order:
#   1. Salutation + name + any supported date   (DD.MM.YYYY or DD-Mon-YYYY)
#   2. Salutation + name + YYYY-MM-DD date
#   3. "Client:" / "Guest:" / "Passenger:" label + name (no date required)
#   4. Salutation + name only (no date at all)
_GUEST_PATTERNS = [
    re.compile(r'^\s*(Mr|Mrs|Ms|Dr|Herr|Frau)\.?\s+(.+?)\s+(' + _DATE_PAT + r')', re.IGNORECASE | re.MULTILINE),
    re.compile(r'^\s*(Mr|Mrs|Ms|Dr|Herr|Frau)\.?\s+(.+?)\s+(\d{4}-\d{2}-\d{2})',   re.IGNORECASE | re.MULTILINE),
    re.compile(r'^(?:client|guest|passenger|travell?er|pax)\s*[:\-]\s*(Mr|Mrs|Ms|Dr|Herr|Frau)?\.?\s*(.+?)$', re.IGNORECASE | re.MULTILINE),
    re.compile(r'^\s*(Mr|Mrs|Ms|Dr|Herr|Frau)\.?\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,4})\s*$', re.IGNORECASE | re.MULTILINE),
]
GUEST_RE = _GUEST_PATTERNS[0]  # kept for compat; _search_guests used internally

# Table row: "description N <date> <date> DayName DayName"
# Supports both DD.MM.YYYY and DD-Mon-YYYY date formats.
TABLE_ROW_RE = re.compile(
    r'^(.*?)\s+(\d{1,2})\s+'
    r'(' + _DATE_PAT + r')\s+(' + _DATE_PAT + r')\s+'
    r'(' + _DAYS + r')\s+(' + _DAYS + r')',
    re.IGNORECASE | re.MULTILINE,
)

# ── Date helpers ───────────────────────────────────────────────────────────────

_DATE_FORMATS = [
    # European / UK / India / Australia — DD first
    '%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y',
    # Textual month — DD Mon YYYY variants
    '%d-%b-%Y', '%d %b %Y', '%d. %b %Y',
    '%d-%B-%Y', '%d %B %Y', '%d. %B %Y',
    # ISO 8601 — YYYY-MM-DD
    '%Y-%m-%d', '%Y/%m/%d',
    # United States — MM/DD/YYYY (tried last to avoid ambiguity)
    '%m/%d/%Y', '%m-%d-%Y',
]

def _parse_date(s: str):
    s = s.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date format: {s!r}")


def _normalise_text(text: str) -> str:
    """Normalise typographic characters that break regex matching."""
    # U+2010 NON-BREAKING HYPHEN → ASCII hyphen
    text = text.replace('\u2010', '-')
    # U+2011 NON-BREAKING HYPHEN, U+2012/2013/2014 dashes → hyphen
    for ch in ('\u2011', '\u2012', '\u2013', '\u2014'):
        text = text.replace(ch, '-')
    return text

def _fmt_date(d) -> str:
    return d.strftime('%d.%m.%Y')

def _day_name(d) -> str:
    return d.strftime('%A')

def _date_range(start_str: str, end_str: str):
    """Yield all dates from start to end inclusive as date objects."""
    cur = _parse_date(start_str)
    end = _parse_date(end_str)
    while cur <= end:
        yield cur
        cur += timedelta(days=1)

# ── Multi-pattern search helpers ───────────────────────────────────────────────

def _search_booking_ref(raw_text: str) -> str:
    """Try every booking-ref pattern and return the first match, or ''."""
    for pat in _BOOKING_REF_PATTERNS:
        m = pat.search(raw_text)
        if m:
            return m.group(1).strip()
    return ''


def _search_guests(raw_text: str) -> list:
    """
    Try each guest pattern in order of specificity.
    Returns a deduplicated list of guest dicts as soon as any pattern finds matches.
    """
    # Patterns 0 & 1 — salutation + name (+ date); group(1)=title, group(2)=name_nat
    for pat in _GUEST_PATTERNS[:2]:
        guests, seen = [], set()
        for m in pat.finditer(raw_text):
            title    = (m.group(1) or '').strip().rstrip('.')
            name_nat = m.group(2).strip()
            first, last = _parse_name_nationality(name_nat)
            key = (title.lower(), last.lower())
            if key not in seen:
                seen.add(key)
                guests.append({'title': title or 'Mr', 'first_name': first, 'last_name': last})
        if guests:
            return guests

    # Pattern 2 — "Client: [Title?] Name"
    guests, seen = [], set()
    for m in _GUEST_PATTERNS[2].finditer(raw_text):
        title    = (m.group(1) or '').strip().rstrip('.')
        name_raw = m.group(2).strip()
        first, last = _parse_name_nationality(name_raw)
        key = (title.lower(), last.lower())
        if key not in seen:
            seen.add(key)
            guests.append({'title': title or 'Mr', 'first_name': first, 'last_name': last})
    if guests:
        return guests

    # Pattern 3 — bare salutation + name, no date
    guests, seen = [], set()
    for m in _GUEST_PATTERNS[3].finditer(raw_text):
        title    = (m.group(1) or '').strip().rstrip('.')
        name_raw = m.group(2).strip()
        first, last = _parse_name_nationality(name_raw)
        key = (title.lower(), last.lower())
        if key not in seen:
            seen.add(key)
            guests.append({'title': title or 'Mr', 'first_name': first, 'last_name': last})
    return guests


# ── Guest name parser ──────────────────────────────────────────────────────────

_TWO_WORD_NATS = {'south african', 'new zealand', 'hong kong', 'saudi arabian'}

def _parse_name_nationality(raw: str) -> tuple[str, str]:
    """
    'First(s) Last Nationality' → (first_names, last_name).
    Drops the trailing nationality word(s), then splits name.
    """
    words = raw.strip().split()
    if len(words) >= 3 and ' '.join(words[-2:]).lower() in _TWO_WORD_NATS:
        words = words[:-2]
    elif len(words) >= 2:
        words = words[:-1]
    if not words:
        return raw, ''
    return ' '.join(words[:-1]), words[-1]

# ── Row classification helpers ─────────────────────────────────────────────────

def _extract_city(desc: str) -> str | None:
    """
    Return the city name if this row is a hotel-listing row, else None.

    Patterns handled:
      "Hotel Shimla: 1. Willow Banks"    → Shimla
      "Delhi: 1 Deventure..."            → Delhi
      "Dharamsala 1: Norling..."         → Dharamsala   (option-N in city ref)
      "Corbett: Hotel Tiger Camp"        → Corbett
    """
    # "Hotel [City]: ..."
    m = re.match(r'^Hotel\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*(?:\s+\d+)?\s*:', desc)
    if m:
        return m.group(1).title()
    # "[City] [N]: [digit or Hotel]"  — standard numbered-option hotel rows
    m = re.match(
        r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*(?:\s+\d+)?\s*:\s*'
        r'(?:\d+[\.\:\s]|Hotel\s)',
        desc, re.IGNORECASE,
    )
    if m:
        return m.group(1).title()
    # "[City] N: [any text]"  — e.g. "Dharamsala 1: Norling Guest House..."
    m = re.match(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s+\d+\s*:', desc)
    if m:
        return m.group(1).title()
    # "[City]: [any hotel/lodge text]"  — e.g. "Chitwan: Sapana Lodge package..."
    # Only matches if the word after the colon looks like a hotel name (title-case word),
    # to avoid false positives on activity descriptions like "Transfer Chitwan - Kathmandu".
    m = re.match(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*:\s*([A-Z][a-z])', desc)
    if m:
        return m.group(1).title()
    # "[City] - [... Hotel / Lodge / Villa]" — dash-separated hotel rows
    m = re.match(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s+-\s+', desc)
    if m and re.search(r'hotel|lodge|villa|resort|inn|guest\s*house|palace', desc, re.IGNORECASE):
        return m.group(1).title()
    return None


def _extract_hotel_name(desc: str) -> str:
    """Extract the first hotel-option name from a hotel-listing description."""
    # Strip city/Hotel prefix up to the colon
    stripped = re.sub(r'^(?:Hotel\s+)?\w+(?:\s+\d+)?\s*(?:\s+\d+)?\s*:\s*', '', desc).strip()
    # Strip leading option number  "1. " / "1 " / "1: "
    stripped = re.sub(r'^\d+[\.\:\s]+', '', stripped).strip()
    # Take up to the next numbered option "2." / "2 "
    m = re.match(r'(.+?)(?:\s+\d+[\.\:])', stripped)
    return m.group(1).strip() if m else stripped.strip()


_SECTION_TRANSPORT = ('flight', 'train', 'transfer', 'drive', 'transport', 'daytrain', 'coach')

def _is_section_header(row: dict) -> bool:
    """
    Section headers span multiple dates and summarise a segment.
    Multi-day transport rows are NOT section headers.
    """
    if row['start_date'] == row['end_date']:
        return False
    return not any(w in row['desc'].lower() for w in _SECTION_TRANSPORT)


_SKIP_PHRASES = ('client information only', 'international flight')

def _clean_activity(desc: str) -> str:
    desc = desc.strip().rstrip('.')
    return (desc[0].upper() + desc[1:] + '.') if desc else ''

# ── Main parser ────────────────────────────────────────────────────────────────

def parse_pdf_rules(raw_text: str) -> dict:

    # ── Normalise typographic characters (e.g. U+2010 non-breaking hyphens) ──
    raw_text = _normalise_text(raw_text)

    # ── DEBUG: set env var CONVERTER_DEBUG_TXT=path to dump extracted PDF text ─
    import os as _os
    _debug_path = _os.environ.get('CONVERTER_DEBUG_TXT', '')
    if _debug_path:
        try:
            with open(_debug_path, 'w', encoding='utf-8') as _f:
                _f.write(raw_text)
        except Exception:
            pass

    # 1. Booking reference — tries multiple label phrasings ────────────────────
    booking_ref = _search_booking_ref(raw_text)

    # 2. Guests — tries multiple formats ──────────────────────────────────────
    guests = _search_guests(raw_text)

    # 3. Parse table rows ──────────────────────────────────────────────────────
    rows = []
    for m in TABLE_ROW_RE.finditer(raw_text):
        desc = m.group(1).strip()
        if not desc:
            continue
        if any(s in desc.lower() for s in _SKIP_PHRASES):
            continue
        # Normalise dates to canonical DD.MM.YYYY so all dict lookups are consistent
        rows.append({
            'desc':       desc,
            'daynum':     int(m.group(2)),
            'start_date': _fmt_date(_parse_date(m.group(3))),
            'end_date':   _fmt_date(_parse_date(m.group(4))),
            'day_from':   m.group(5),
        })

    # 4. Index rows by start date ──────────────────────────────────────────────
    date_rows: dict[str, list] = defaultdict(list)
    for row in rows:
        date_rows[row['start_date']].append(row)

    # 5. Build hotel_by_date ───────────────────────────────────────────────────
    # Two passes so that a hotel starting today always wins over one continuing.
    hotel_by_date: dict[str, tuple] = {}

    # Pass A: fill continuation dates (lower priority — first-come wins)
    for row in rows:
        city = _extract_city(row['desc'])
        if not city:
            continue
        hotel_name = _extract_hotel_name(row['desc'])
        for d in _date_range(row['start_date'], row['end_date']):
            ds = _fmt_date(d)
            if ds not in hotel_by_date:
                hotel_by_date[ds] = (city, hotel_name)

    # Pass B: overwrite with the hotel that actually starts on each date
    for row in rows:
        city = _extract_city(row['desc'])
        if city:
            hotel_by_date[row['start_date']] = (city, _extract_hotel_name(row['desc']))

    # 6. Determine full date range (arrival → departure) ───────────────────────
    # Use the earliest and latest dates that appear in any row.
    all_row_dates = [_parse_date(r['start_date']) for r in rows] + \
                    [_parse_date(r['end_date'])   for r in rows]
    if not all_row_dates:
        return {'quotation': {}, 'guests': guests, 'itinerary': [], 'parsing_success': False, 'parsing_errors': ['No itinerary rows found — the PDF table format may not be recognised.']}

    # Arrival = first date with actual rows (not just hotel coverage)
    # Departure = last row start date
    arrival_date   = min(_parse_date(d) for d in date_rows)
    departure_date = max(_parse_date(d) for d in date_rows)
    nights         = (departure_date - arrival_date).days

    # Build the complete date sequence including leisure days with no rows
    full_range = list(_date_range(_fmt_date(arrival_date), _fmt_date(departure_date)))

    # 7. Build itinerary ───────────────────────────────────────────────────────
    itinerary = []
    for i, d in enumerate(full_range):
        date_str = _fmt_date(d)
        is_first = (i == 0)
        is_last  = (i == len(full_range) - 1)

        city, hotel_name = hotel_by_date.get(date_str, (None, None))
        day_rows = date_rows.get(date_str, [])

        # Fallback city from non-hotel activity description.
        # Skip section-header rows and exclude common non-city lead words.
        _NON_CITY = {
            'transfer', 'private', 'half', 'afternoon', 'the', 'in', 'on',
            'intercity', 'daytrain', 'train', 'flight', 'module', 'overnight',
            'kings', 'gods', 'coach', 'drive', 'day', 'full', 'early', 'late',
        }
        if not city:
            for row in day_rows:
                if _is_section_header(row):
                    continue
                m2 = re.match(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s', row['desc'])
                if m2 and m2.group(1).lower() not in _NON_CITY:
                    city = m2.group(1).title()
                    break

        # Does a hotel START today?
        hotel_starts_today = any(
            _extract_city(r['desc']) and r['start_date'] == date_str
            for r in day_rows
        )

        # ── Build description ─────────────────────────────────────────────────
        parts = []

        if is_first:
            parts.append(f"Arrive in {city or 'destination'} by international flight.")

        if not day_rows:
            # Leisure / rest day — no entries in the PDF
            parts.append(f"Day at leisure in {city or 'destination'}.")
        else:
            for row in day_rows:
                if _extract_city(row['desc']):    # hotel listing → city already captured
                    continue
                if _is_section_header(row):        # segment summary → skip
                    continue
                parts.append(_clean_activity(row['desc']))

        if hotel_starts_today and hotel_name:
            parts.append(f"Stay at {hotel_name}.")

        if is_last:
            parts.append("Transfer to airport for international flight home.")

        description = ' '.join(p for p in parts if p) or f"Day in {city or 'transit'}."

        itinerary.append({
            'sr_no':       i + 1,
            'day_name':    _day_name(d),
            'date':        date_str,
            'destination': city or 'In Transit',
            'description': description,
        })

    # 8. Quotation fields ──────────────────────────────────────────────────────
    num_pax   = len(guests)
    principal = guests[0] if guests else {}
    principal_name = (
        f"{principal.get('title', '')}. "
        f"{principal.get('first_name', '')} "
        f"{principal.get('last_name', '')}"
    ).strip('. ')

    raw_lower = raw_text.lower()

    quotation = {
        'QuotationRef':    booking_ref,
        'PaxName':         principal.get('last_name', ''),
        'PaxFirstName':    principal.get('first_name', ''),
        'PrincipalClient': principal_name,
        'Email':           None,
        'NumPax':          num_pax,
        'NumSingles':      1 if num_pax == 1 else 0,
        'NumDoubles':      1 if num_pax == 2 else 0,
        'NumTwins':        0,
        'NumTriples':      1 if num_pax == 3 else 0,
        'DateOfArrival':   _fmt_date(arrival_date),
        'DateOfDeparture': _fmt_date(departure_date),
        'StartDate':       _fmt_date(arrival_date),
        'EndDate':         _fmt_date(departure_date),
        'Nights':          nights,
        'FlightNo':        None,
        'FlightNoDept':    None,
        'PlaceFrom':       None,
        'PlaceTo':         None,
        'ETA':             None,
        'ETD':             None,
        'Guide':           'guide' in raw_lower,
        'EntranceFees':    'entrance fee' in raw_lower,
        'TourCode':        None,
        'Comment':         None,
    }

    # ── Validate required fields ───────────────────────────────────────────────
    parsing_errors = []
    if not booking_ref:
        parsing_errors.append(
            "Booking reference not found — check that the PDF contains a label "
            "such as 'Booking No', 'Booking Number', or 'Ref'."
        )
    if not guests:
        parsing_errors.append(
            "Client name not found — check that the PDF lists passengers with a "
            "salutation (Mr/Mrs/Ms/Dr) or a 'Client:' / 'Guest:' label."
        )
    elif not principal_name.strip('. '):
        parsing_errors.append("Client name could not be parsed from the guest list.")

    parsing_success = len(parsing_errors) == 0

    return {
        'quotation':       quotation,
        'guests':          guests,
        'itinerary':       itinerary,
        'parsing_success': parsing_success,
        'parsing_errors':  parsing_errors,
    }


# ── Public API ─────────────────────────────────────────────────────────────────

def extract_pdf_text(pdf_path: str) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        pages = [p.extract_text(layout=True) for p in pdf.pages]
    return '\n\n--- PAGE BREAK ---\n\n'.join(t for t in pages if t)


def parse_with_claude(raw_text: str) -> dict:
    """Name kept for app.py compatibility — calls local rule-based parser."""
    return parse_pdf_rules(raw_text)


# ──────────────────────────────────────────────────────────────
# EXCEL BUILDING WITH PROFESSIONAL FORMATTING
# ──────────────────────────────────────────────────────────────

def to_date(date_str):
    """Convert DD.MM.YYYY string to date object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%d.%m.%Y").date()
    except (ValueError, TypeError):
        return None

def guest_display(guests: list) -> str:
    """Format guest list as readable string."""
    if not guests:
        return "Guest"
    return " & ".join(f"{g.get('title', 'Mr')}. {g.get('first_name', '')} {g.get('last_name', '')}" for g in guests)

# Style constants
THIN = Side(style="thin", color="BFBFBF")
def tb():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BLUE      = "1F4E79"
ALT_BLUE  = "DCE6F1"
GREEN     = "375623"
ALT_GREEN = "EBF1DE"

def hdr_style(color_hex):
    return (
        Font(name="Calibri", bold=True, color="FFFFFF", size=10),
        PatternFill("solid", fgColor=color_hex),
    )

DAT_FONT = Font(name="Calibri", size=10)
LBL_FONT = Font(name="Calibri", bold=True, size=10)

# DB columns with friendly labels
DB_COLUMNS = [
    ("QuotationRef",    "Quotation Ref",        "QuotationRef",    "text"),
    ("PrincipalClient", "Principal Client",     "PrincipalClient", "text"),
    ("PaxName",         "Pax Last Name",        "PaxName",         "text"),
    ("PaxFirstName",    "Pax First Name(s)",    "PaxFirstName",    "text"),
    ("Email",           "Email",                "Email",           "text"),
    ("NumPax",          "No. of Pax",           "NumPax",          "int"),
    ("NumSingles",      "Single Rooms",         "NumSingles",      "int"),
    ("NumDoubles",      "Double Rooms",         "NumDoubles",      "int"),
    ("NumTwins",        "Twin Rooms",           "NumTwins",        "int"),
    ("NumTriples",      "Triple Rooms",         "NumTriples",      "int"),
    ("DateOfArrival",   "Date of Arrival",      "DateOfArrival",   "date"),
    ("DateOfDeparture", "Date of Departure",    "DateOfDeparture", "date"),
    ("StartDate",       "Start Date",           "StartDate",       "date"),
    ("EndDate",         "End Date",             "EndDate",         "date"),
    ("Nights",          "No. of Nights",        "Nights",          "int"),
    ("FlightNo",        "Inbound Flight No.",   "FlightNo",        "text"),
    ("ETA",             "ETA",                  "ETA",             "text"),
    ("PlaceFrom",       "Place From",           "PlaceFrom",       "text"),
    ("FlightNoDept",    "Outbound Flight No.",  "FlightNoDept",    "text"),
    ("ETD",             "ETD",                  "ETD",             "text"),
    ("PlaceTo",         "Place To",             "PlaceTo",         "text"),
    ("Guide",           "Guide Included",       "Guide",           "bool"),
    ("EntranceFees",    "Entrance Fees Incl.",  "EntranceFees",    "bool"),
    ("Comment",         "Comment / Notes",      "Comment",         "text"),
]

def build_db_sheet(ws, q: dict):
    """Build the DB Import sheet with friendly labels."""
    ws.title = "DB Import"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Quotations — DB Import"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    hdr_font, hdr_fill = hdr_style(BLUE)
    for col, label in enumerate(["Friendly Label", "DB Column Name", "Value"], start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb()
    ws.row_dimensions[2].height = 18

    # Data rows
    for r, (db_col, label, key, dtype) in enumerate(DB_COLUMNS, start=3):
        raw = q.get(key)

        if   dtype == "date": val = to_date(raw)
        elif dtype == "int":  val = int(raw) if raw is not None else 0
        elif dtype == "bool": val = "Yes" if raw else "No"
        else:                 val = raw if raw is not None else ""

        alt = PatternFill("solid", fgColor=ALT_BLUE) if r % 2 == 0 else None

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LBL_FONT
        lc.border = tb()
        lc.alignment = Alignment(vertical="center")
        if alt: lc.fill = alt

        dc = ws.cell(row=r, column=2, value=db_col)
        dc.font = Font(name="Courier New", size=9, color="555555")
        dc.border = tb()
        dc.alignment = Alignment(vertical="center")
        if alt: dc.fill = alt

        vc = ws.cell(row=r, column=3, value=val)
        vc.font = DAT_FONT
        vc.border = tb()
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        if dtype == "date" and val:
            vc.number_format = "DD/MM/YYYY"
        if alt: vc.fill = alt

        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A3"


def build_itinerary_sheet(ws, data: dict):
    """Build the Itinerary sheet with day-by-day details."""
    ws.title = "Itinerary"
    for col, w in zip("ABCDE", [8, 12, 14, 20, 80]):
        ws.column_dimensions[col].width = w

    # Title (guest names)
    ws.merge_cells("A1:E1")
    guests = data.get("guests", [])
    ws["A1"] = guest_display(guests)
    ws["A1"].font = Font(name="Calibri", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Summary row
    q = data.get("quotation", {})
    parts = [f"{q.get(k, 0)}×{lbl}" for k, lbl in
             [("NumDoubles","Double"),("NumTwins","Twin"),("NumSingles","Single"),("NumTriples","Triple")]
             if q.get(k)]
    ws.merge_cells("A2:E2")
    summary = f"Pax: {q.get('NumPax', 0)}   |   Rooms: {', '.join(parts)}   |   Nights: {q.get('Nights', 0)}"
    ws["A2"] = summary
    ws["A2"].font = Font(name="Calibri", bold=True, size=10, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Column headers
    hdr_font, hdr_fill = hdr_style(GREEN)
    for col, hdr in enumerate(["Sr. No", "Day", "Date", "Destination", "Itinerary"], start=1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb()
    ws.row_dimensions[4].height = 18

    # Day rows
    for idx, entry in enumerate(data.get("itinerary", [])):
        row = 5 + idx
        alt = PatternFill("solid", fgColor=ALT_GREEN) if idx % 2 == 1 else None
        vals = [entry.get("sr_no"), entry.get("day_name"), to_date(entry.get("date")),
                entry.get("destination"), entry.get("description")]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = DAT_FONT
            c.border = tb()
            if alt: c.fill = alt
            if col in (1, 2, 4):
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif col == 3:
                c.number_format = "DD/MM/YYYY"
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 44

    ws.freeze_panes = "A5"


def build_excel(data: dict, output_path: str):
    """Build a professional two-sheet Excel workbook."""
    wb = Workbook()
    build_db_sheet(wb.active, data.get("quotation", {}))
    build_itinerary_sheet(wb.create_sheet(), data)
    wb.save(output_path)


def dump_pdf_text(pdf_path: str, out_path: str = None) -> str:
    """
    Extract raw text from a PDF and print the first 3000 chars to stdout.
    Also saves the full text to out_path if given.
    Call this from a one-off script to see exactly what pdfplumber extracts
    — paste the output so patterns can be fixed.

    Usage:
        python - <<'EOF'
        from converter import dump_pdf_text
        dump_pdf_text("uploads/booking_XYZ/booking_XYZ.pdf", "debug_text.txt")
        EOF
    """
    text = extract_pdf_text(pdf_path)
    print("=" * 60)
    print("PDF RAW TEXT (first 3000 chars)")
    print("=" * 60)
    print(text[:3000])
    print("=" * 60)
    if out_path:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"Full text saved to: {out_path}")
    return text


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python converter.py <path/to/file.pdf> [output_text.txt]")
        sys.exit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else None
    dump_pdf_text(sys.argv[1], out)